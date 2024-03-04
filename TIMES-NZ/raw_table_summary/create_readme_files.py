"""
Script to auto-generate README files for the structured workbook documentation, using the information in `raw_tables.txt`.

Behaviour:
- Deletes README files that are not relevant according to the information in `raw_tables.txt`.
- Creates README files for workbooks and sheets that do not already exist.
- Leaves existing README files unchanged.

Usage:
`python create_readme_files.py`
"""


import os
import shutil
import openpyxl

#### CONSTANTS ####
FILEPATH = 'raw_tables.txt'
BASE_DIR = '../../docs'
OVERWRITE_READMES = True

#### FUNCTIONS ####
def parse_raw_tables_file(filepath):
    """
    Parse the raw_tables.txt file to extract workbook and sheet information.

    Parameters:
    - filepath: Path to the raw_tables.txt file.

    Returns:
    - A dictionary with workbook names as keys and sheet information as values.
    """
    with open(filepath, 'r', encoding='utf-8') as file:
        text = file.read()
    blocks = text.strip().split('\n\n')
    parsed_data = {}
    for block in blocks:
        lines = block.strip().split('\n')
        if not lines:
            continue
        block_info = {
            'sheetname': None,
            'range': None,
            'filename': None,
            'tag': None,
            'types': []
        }
        for line in lines:
            if ': ' in line:
                key, value = line.split(': ', 1)  # Split only on the first occurrence
                if key == 'sheetname':
                    block_info['sheetname'] = value
                elif key == 'range':
                    block_info['range'] = value
                elif key == 'filename':
                    block_info['filename'] = value
                elif key == 'tag':
                    block_info['tag'] = value
                elif key == 'types':
                    types_list = [type_.split(' ')[0] for type_ in value.split(', ')]
                    block_info['types'] = types_list
            else:
                pass
        if block_info['filename'] not in parsed_data:
            parsed_data[block_info['filename']] = {}
        if block_info['sheetname'] not in parsed_data[block_info['filename']]:
            parsed_data[block_info['filename']][block_info['sheetname']] = []
        parsed_data[block_info['filename']][block_info['sheetname']].append(block_info)
    return parsed_data


def read_documentation_from_excel(workbook_path):
    """
    Read documentation from the "Documentation" sheet in an Excel workbook.
    Assumes that documentation in the "Documentation" sheet is in the format:
    - "Workbook: <documentation for the workbook>"
    - "<sheet name>: <documentation for the sheet>"
    - ...

    Parameters:
    - workbook_path: Path to the Excel workbook.

    Returns:
    - A dictionary with two keys: 'workbook' for workbook-level documentation,
      and 'sheets' for sheet-specific documentation.
    """
    documentation = {'workbook': '', 'sheets': {}}
    try:
        print(workbook_path)
        workbook = openpyxl.load_workbook(workbook_path, data_only=True)
        if 'Documentation' in workbook.sheetnames:
            sheet = workbook['Documentation']
            for row in sheet.iter_rows(min_row=1, values_only=True):
                if row[0] and isinstance(row[0], str):
                    if row[0].startswith('Workbook'):
                        documentation['workbook'] = row[0].replace('Workbook: ', '', 1).strip()
                    else:
                        sheet_name = row[0].split(':')[0].strip()
                        doc_text = row[0].split(':', 1)[1].strip() if ':' in row[0] else ''
                        documentation['sheets'][sheet_name] = doc_text
    except Exception as e:
        print(f"Error reading documentation from {workbook_path}: {e}")
    return documentation


def prune_directory_tree(expected_structure, actual_base_dir, test_run=False):
    """
    Prune the directory tree to match the expected structure.

    Parameters:
    - expected_structure: A dictionary representing the expected directories and files.
    - actual_base_dir: The base directory of the actual structure to be pruned.
    """
    expected_paths = set()
    for directory, files in expected_structure.items():
        expected_paths.add(directory)
        for file in files:
            expected_paths.add(os.path.join(directory, file))
    actual_paths = set()
    for root, dirs, files in os.walk(actual_base_dir):
        for name in files:
            actual_path = os.path.relpath(os.path.join(root, name), actual_base_dir)
            actual_paths.add(actual_path)
        for name in dirs:
            actual_path = os.path.relpath(os.path.join(root, name), actual_base_dir)
            actual_paths.add(actual_path)
    paths_to_prune = actual_paths - expected_paths
    if test_run:
        print(paths_to_prune)
        return
    for path in paths_to_prune:
        full_path = os.path.join(actual_base_dir, path)
        if os.path.isfile(full_path):
            print("Removing file:", full_path)
            os.remove(full_path)
        elif os.path.isdir(full_path):
            print("Removing directory:", full_path)
            shutil.rmtree(full_path, ignore_errors=True)


def intermediate_dirs(directory):
    intermediate_dirs = []
    parts = os.path.normpath(directory).split(os.sep)
    for i in range(1, len(parts) + 1):
        intermediate_dirs.append(os.path.join(*parts[:i]))
    return intermediate_dirs


def expected_directory_structure(parsed_data):
    """
    Generate the expected directory structure based on parsed data.

    Parameters:
    - parsed_data: Parsed data from the raw_tables.txt file.

    Returns:
    - A dictionary representing the expected directories and files.
    """
    expected_structure = {}
    for workbook, sheets in parsed_data.items():
        if not workbook:
            continue
        workbook_dir = os.path.normpath(os.path.splitext(workbook)[0])
        for directory in intermediate_dirs(workbook_dir):
            expected_structure[directory] = []
        expected_structure[workbook_dir] = ['README.md']
        for sheet in sheets:
            if not sheet:
                continue
            # Add sheet README.md to expected structure
            expected_structure[workbook_dir].append(f"{sheet}.md")
    return expected_structure


def create_readme_files(parsed_data, base_dir):
    """
    Create README.md files for each workbook and sheet based on parsed data.

    Parameters:
    - parsed_data: Parsed data from the raw_tables.txt file.
    - base_dir: Base directory where the README files will be created.
    """
    base_dir = os.path.normpath(base_dir)
    os.makedirs(base_dir, exist_ok=True)
    if OVERWRITE_READMES:
        prune_directory_tree({}, base_dir)
    else:
        expected_structure = expected_directory_structure(parsed_data)
        prune_directory_tree(expected_structure, base_dir)
    index_readme_path = os.path.join(base_dir, 'README.md')
    with open(index_readme_path, 'w', encoding='utf-8') as index_readme:
        index_readme_content = f"[Back to Main Documentation](../README.md)\n\n"
        index_readme_content += "# Structured Workbook Documentation\n\n## Workbooks Index\n\n"
        for workbook, _ in parsed_data.items():
            if not workbook:
                continue
            workbook_name = os.path.basename(workbook)
            workbook_path_without_ext = os.path.normpath(os.path.splitext(workbook)[0])
            workbook_dir_path = os.path.dirname(os.path.normpath(os.path.splitext(workbook)[0])).replace('\\', '/')
            if not workbook_dir_path:
                workbook_dir_path = '.'
            workbook_rel_path = os.path.join(workbook_path_without_ext, 'README.md').replace('\\', '/')
            index_readme_content += f"- `{workbook_dir_path}/`[{workbook_name}]({workbook_rel_path})\n"
        print(f'Write index {index_readme_path}')
        index_readme.write(index_readme_content)
    for workbook, sheets in parsed_data.items():
        if not workbook:
            continue
        workbook_name = os.path.basename(workbook)
        workbook_dir = os.path.normpath(os.path.join(base_dir, os.path.splitext(workbook)[0]))
        os.makedirs(workbook_dir, exist_ok=True)  # Ensure workbook directory exists
        depth = workbook_dir.count(os.sep) - base_dir.count(os.sep)
        back_to_index = "../" * depth + "README.md"
        workbook_readme_path = os.path.join(workbook_dir, 'README.md')

        xlsx_relpath = os.path.join('..', os.path.normpath(workbook)).replace('\\', '/')
        documentation = read_documentation_from_excel(xlsx_relpath)
        workbook_doc = documentation.get('workbook', '')
        if not os.path.exists(workbook_readme_path):
            with open(workbook_readme_path, 'w', encoding='utf-8') as workbook_readme:
                workbook_readme_content = f"# {os.path.basename(workbook)}\n\n"
                workbook_readme_content += f"[Back to Index]({back_to_index})\n\n"
                workbook_readme_content += "## Workbook Overview\n\n"
                workbook_readme_content += f"{workbook_doc}\n\n" if workbook_doc \
                    else "(TODO: Add a high-level overview of how this workbook fits into the TIMES-NZ model.)\n\n"
                for sheet in sheets:
                    sheet_readme_filename = f"{sheet}.md"
                    workbook_readme_content += f"- [{sheet}]({sheet_readme_filename.replace(' ', '%20')}) - Overview of the '{sheet}' sheet.\n"
                print(f'Write workbook documentation {workbook_readme_path}')
                workbook_readme.write(workbook_readme_content)
        for sheet, info in sheets.items():
            if not sheet:
                continue
            sheet_readme_filename = f"{sheet}.md"
            sheet_readme_path = os.path.join(workbook_dir, sheet_readme_filename)
            sheet_doc = documentation['sheets'].get(sheet, '')
            if not os.path.exists(sheet_readme_path):
                with open(sheet_readme_path, 'w', encoding='utf-8') as sheet_readme:
                    sheet_readme_content = f"# {sheet}\n\n"
                    sheet_readme_content = f"[Back to {workbook_name}](README.md)\n\n# '{sheet}' sheet in {workbook_name}\n\n"
                    sheet_readme_content += "## Sheet Overview\n\n"
                    sheet_readme_content += f"{sheet_doc}\n\n" if sheet_doc \
                        else "(TODO: Overview of the sheet. Units used, sources of data, etc.)\n\n"
                    for entry in info:
                        sheet_readme_content += f"#### Table definition: {entry['tag']}\n"
                        sheet_readme_content += f"- **Range**: {entry['range']}\n"
                        if entry['types']:
                            fixed = [x.replace('~','\~') for x in entry['types']]
                            sheet_readme_content += "- **Columns**: '" + "', '".join(fixed) + "'\n\n"
                    print(f'Write sheet documentation {sheet_readme_path}')
                    sheet_readme.write(sheet_readme_content)


#### MAIN ####
parsed_data = parse_raw_tables_file(FILEPATH)
create_readme_files(parsed_data, BASE_DIR)